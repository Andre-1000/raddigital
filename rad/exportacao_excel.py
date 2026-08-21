"""
Exportacao para Excel (.xlsx) -- 21/08/2026.

Formato "1 RAD = 1 linha, 1 campo = 1 coluna", pensado para consumo em
ferramentas de BI (Excel/Power Query, Looker Studio etc.) usadas pelos
colaboradores para acompanhar as pesquisas diarias -- nao para o
layout oficial do documento RDA (esse continua em
rad/exportacao_oficial.py, formato Word).

Campos de multipla selecao (Linhas, Vias, Equipes, Servicos,
Colaboradores, MCHs do bloco AMV) sao achatados em uma unica celula de
texto, separados por '; ' -- nao ha como representar uma relacao
1-para-N numa planilha onde cada RAD e uma linha so. Quem precisar do
detalhe individual (ex.: por MCH ou por colaborador) deve usar a tela
de Consulta ou a API, nao o Excel.
"""
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COLUNAS = [
    ('numero_rad', 'Número RAD'),
    ('numero_os', 'OS'),
    ('numero_sa', 'SA'),
    ('status', 'Status'),
    ('data_preenchimento', 'Data'),
    ('local_inicial', 'Local Inicial'),
    ('local_final', 'Local Final'),
    ('tipo_manutencao', 'Tipo de Manutenção'),
    ('numero_falha', 'Nº Falha'),
    ('hora_prog_inicio', 'Hora Prog. Início'),
    ('hora_prog_termino', 'Hora Prog. Término'),
    ('hora_real_inicio', 'Hora Real Início'),
    ('hora_real_termino', 'Hora Real Término'),
    ('duracao_programada_min', 'Duração Programada (min)'),
    ('duracao_real_min', 'Duração Real (min)'),
    ('atraso_inicio', 'Atraso Início'),
    ('motivo_atraso_inicio', 'Motivo Atraso Início'),
    ('atraso_termino', 'Atraso Término'),
    ('motivo_atraso_termino', 'Motivo Atraso Término'),
    ('linhas', 'Linhas'),
    ('vias', 'Vias'),
    ('equipes', 'Equipes'),
    ('servicos', 'Serviços'),
    ('mchs', 'MCHs (Bloco AMV)'),
    ('colaboradores', 'Colaboradores'),
    ('login_usuario', 'Login de Quem Preencheu'),
    ('dispositivo', 'Dispositivo'),
    ('data_sincronizacao', 'Data de Sincronização'),
]


def _valor_ou_vazio(valor):
    return valor if valor not in (None, '') else ''


def _linha_para_rad(rad):
    return {
        'numero_rad': rad.numero_rad,
        'numero_os': rad.numero_os,
        'numero_sa': rad.numero_sa,
        'status': rad.get_status_display(),
        'data_preenchimento': rad.data_preenchimento,
        'local_inicial': rad.local_inicial.sigla,
        'local_final': rad.local_final.sigla,
        'tipo_manutencao': rad.tipo_manutencao.nome,
        'numero_falha': rad.numero_falha,
        'hora_prog_inicio': rad.hora_prog_inicio,
        'hora_prog_termino': rad.hora_prog_termino,
        'hora_real_inicio': rad.hora_real_inicio,
        'hora_real_termino': rad.hora_real_termino,
        'duracao_programada_min': rad.duracao_programada_min,
        'duracao_real_min': rad.duracao_real_min,
        'atraso_inicio': 'Sim' if rad.atraso_inicio else 'Não',
        'motivo_atraso_inicio': (
            rad.motivo_atraso_inicio.nome if rad.motivo_atraso_inicio else ''
        ),
        'atraso_termino': 'Sim' if rad.atraso_termino else 'Não',
        'motivo_atraso_termino': (
            rad.motivo_atraso_termino.nome if rad.motivo_atraso_termino else ''
        ),
        'linhas': '; '.join(str(v) for v in rad.linhas.values_list('linha_id', flat=True)),
        'vias': '; '.join(rad.vias.values_list('via__nome', flat=True)),
        'equipes': '; '.join(rad.equipes.values_list('equipe_id', flat=True)),
        'servicos': '; '.join(rad.servicos.values_list('servico__nome', flat=True)),
        'mchs': '; '.join(amv.mch.identificacao for amv in rad.amv_blocos.all()),
        'colaboradores': '; '.join(c.nome for c in rad.colaboradores.all()),
        'login_usuario': rad.usuario_id,
        'dispositivo': rad.get_dispositivo_display(),
        'data_sincronizacao': rad.data_sincronizacao,
    }


def gerar_excel_bytes(rads):
    """
    Recebe uma lista (ou queryset ja avaliado) de Rad, com
    select_related/prefetch_related ja aplicados pelo chamador, e
    retorna os bytes de um .xlsx pronto para download -- uma linha por
    RAD.
    """
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = 'RADs'

    cabecalho = [rotulo for _, rotulo in COLUNAS]
    planilha.append(cabecalho)

    for rad in rads:
        dados = _linha_para_rad(rad)
        planilha.append([_valor_ou_vazio(dados.get(chave)) for chave, _ in COLUNAS])

    for indice, (_, rotulo) in enumerate(COLUNAS, start=1):
        largura = max(len(rotulo) + 2, 14)
        planilha.column_dimensions[get_column_letter(indice)].width = largura

    buffer_saida = io.BytesIO()
    workbook.save(buffer_saida)
    return buffer_saida.getvalue()
